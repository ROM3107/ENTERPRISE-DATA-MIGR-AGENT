#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define color scheme
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
ALT_ROW_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
SUMMARY_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Status colors
STATUS_READY = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
STATUS_PENDING = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
STATUS_REFACTOR = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
STATUS_LOW = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
STATUS_MEDIUM = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
STATUS_HIGH = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")

# Font styling
STATUS_FONT = Font(bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(size=10)
WHITE_FONT = Font(color="FFFFFF")

# Border styling
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ===== SUMMARY SHEET =====
ws_summary = wb.create_sheet('Summary', 0)
summary_data = [
    ['Metric', 'Value'],
    ['Total Components', '7'],
    ['Functions', '3'],
    ['Procedures', '4'],
    ['Ready for Migration', '5'],
    ['Requires Refactoring', '2'],
    ['Oracle-Specific Features', '5'],
    ['Data Types to Map', '4'],
    ['Estimated Total Effort', '10.5 hours'],
    ['Overall Complexity', 'Medium'],
    ['Migration Assessment', 'FEASIBLE - PROCEED']
]

for idx, row in enumerate(summary_data):
    ws_summary.append(row)
    if idx == 0:  # Header row
        for cell in ws_summary[1]:
            cell.fill = SUMMARY_HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER
    else:
        # Alternate row colors
        fill_color = ALT_ROW_FILL if idx % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for cell in ws_summary[idx + 1]:
            cell.fill = fill_color
            cell.font = CELL_FONT
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = THIN_BORDER

ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 30
ws_summary.row_dimensions[1].height = 25

# ===== DETAILS SHEET =====
ws_details = wb.create_sheet('Details')
details_headers = ['Component Name', 'Type', 'Status', 'Complexity', 'Effort (hrs)', 'Dependencies', 'Issues']
ws_details.append(details_headers)

# Format header row
for cell in ws_details[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

details_data = [
    ['new_customer', 'Function', 'Ready', 'Medium', 2, 'Sequence setup', 'None'],
    ['get_customer', 'Function', 'Ready', 'Low', 1, 'None', 'None'],
    ['get_customer_name', 'Function', 'Ready', 'Low', 0.5, 'None', 'None'],
    ['set_customer (v1)', 'Procedure', 'Ready', 'Low', 0.5, 'None', 'None'],
    ['set_customer (v2)', 'Procedure', 'Requires Refactoring', 'Medium', 2, 'Overload resolution', 'Overloading not supported'],
    ['delete_customer', 'Procedure', 'Ready', 'Low', 0.5, 'None', 'None'],
    ['purge_old_customers', 'Procedure', 'Requires Specification', 'Medium', 2.5, 'Audit table definition', 'Audit trail logic not implemented']
]

for idx, row in enumerate(details_data):
    ws_details.append(row)
    row_num = idx + 2
    
    # Status-based coloring
    status_cell = ws_details[f'C{row_num}']
    status = row[2]
    if status == 'Ready':
        status_cell.fill = STATUS_READY
    elif status == 'Requires Refactoring' or status == 'Requires Specification':
        status_cell.fill = STATUS_REFACTOR
    status_cell.font = STATUS_FONT
    
    # Complexity-based coloring
    complexity_cell = ws_details[f'D{row_num}']
    complexity = row[3]
    if complexity == 'Low':
        complexity_cell.fill = STATUS_LOW
    elif complexity == 'Medium':
        complexity_cell.fill = STATUS_MEDIUM
    complexity_cell.font = STATUS_FONT
    
    # Apply formatting to all cells in row
    for col_idx, cell_value in enumerate(row):
        cell = ws_details.cell(row=row_num, column=col_idx + 1)
        cell.font = CELL_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        # Alternate row colors for other columns
        if col_idx not in [2, 3]:  # Don't override status and complexity colors
            fill_color = ALT_ROW_FILL if idx % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.fill = fill_color

# Set column widths
ws_details.column_dimensions['A'].width = 20
ws_details.column_dimensions['B'].width = 12
ws_details.column_dimensions['C'].width = 22
ws_details.column_dimensions['D'].width = 15
ws_details.column_dimensions['E'].width = 12
ws_details.column_dimensions['F'].width = 20
ws_details.column_dimensions['G'].width = 30
ws_details.row_dimensions[1].height = 25

# ===== MAPPINGS SHEET =====
ws_mappings = wb.create_sheet('Mappings')
mappings_headers = ['Oracle Feature', 'Snowflake Equivalent', 'Conversion Strategy', 'Risk Level']
ws_mappings.append(mappings_headers)

# Format header row
for cell in ws_mappings[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

mappings_data = [
    ['RETURNING Clause', 'Sequence or MAX() function', 'Use NEXTVAL() with explicit insert', 'Low'],
    ['%type Declaration', 'Explicit type declaration', 'Define types explicitly', 'Low'],
    ['%rowtype Declaration', 'OBJECT or VARIANT type', 'Use OBJECT_CONSTRUCT()', 'Medium'],
    ['Procedure Overloading', 'Naming convention', 'Rename procedures with suffix', 'Medium'],
    ['NO_DATA_FOUND Exception', 'NULL check or LIMIT 1', 'Replace exception with NULL logic', 'Low'],
    ['BOOLEAN Data Type', 'BOOLEAN type', 'Direct mapping', 'Low'],
    ['NUMBER Type', 'BIGINT for IDs', 'Map to BIGINT', 'Low'],
    ['VARCHAR2 Type', 'VARCHAR type', 'Direct mapping', 'Low'],
    ['DATE Type', 'TIMESTAMP_NTZ', 'Direct mapping', 'Low']
]

for idx, row in enumerate(mappings_data):
    ws_mappings.append(row)
    row_num = idx + 2
    
    # Risk level coloring
    risk_cell = ws_mappings[f'D{row_num}']
    risk = row[3]
    if risk == 'Low':
        risk_cell.fill = STATUS_LOW
    elif risk == 'Medium':
        risk_cell.fill = STATUS_MEDIUM
    risk_cell.font = STATUS_FONT
    
    # Apply formatting to all cells
    for col_idx, cell_value in enumerate(row):
        cell = ws_mappings.cell(row=row_num, column=col_idx + 1)
        cell.font = CELL_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        if col_idx != 3:  # Don't override risk level color
            fill_color = ALT_ROW_FILL if idx % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.fill = fill_color

# Set column widths
ws_mappings.column_dimensions['A'].width = 25
ws_mappings.column_dimensions['B'].width = 28
ws_mappings.column_dimensions['C'].width = 35
ws_mappings.column_dimensions['D'].width = 15
ws_mappings.row_dimensions[1].height = 25

# ===== ISSUES SHEET =====
ws_issues = wb.create_sheet('Issues')
issues_headers = ['Issue ID', 'Issue Title', 'Component', 'Severity', 'Description', 'Recommended Solution', 'Resolution Status']
ws_issues.append(issues_headers)

# Format header row
for cell in ws_issues[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

issues_data = [
    ['ISS-001', 'Procedure Overloading Not Supported', 'set_customer', 'High', 'Snowflake does not support procedure overloading with same name but different signatures', 'Use naming convention: set_customer, set_customer_row', 'Pending'],
    ['ISS-002', 'Audit Trail Logic Not Implemented', 'purge_old_customers', 'Medium', 'TODO comment indicates audit trail deletion is not yet implemented', 'Define audit table schema and implement deletion logic', 'Pending'],
    ['ISS-003', 'RETURNING Clause Limited Support', 'new_customer', 'Medium', 'Snowflake has limited support for RETURNING clause', 'Implement sequence-based approach with NEXTVAL()', 'Pending'],
    ['ISS-004', '%rowtype Not Directly Supported', 'get_customer, set_customer', 'Low', 'Snowflake does not have %rowtype anchor type', 'Define OBJECT type or use column-by-column approach', 'Pending'],
    ['ISS-005', 'No Data Found Exception Handling', 'get_customer, get_customer_name', 'Low', 'Oracle exception handling not available in Snowflake', 'Replace exception blocks with NULL value returns', 'Pending']
]

for idx, row in enumerate(issues_data):
    ws_issues.append(row)
    row_num = idx + 2
    
    # Severity coloring
    severity_cell = ws_issues[f'D{row_num}']
    severity = row[3]
    if severity == 'High':
        severity_cell.fill = STATUS_HIGH
    elif severity == 'Medium':
        severity_cell.fill = STATUS_MEDIUM
    elif severity == 'Low':
        severity_cell.fill = STATUS_LOW
    severity_cell.font = STATUS_FONT
    
    # Status coloring
    status_cell = ws_issues[f'G{row_num}']
    status_cell.fill = STATUS_PENDING
    status_cell.font = STATUS_FONT
    
    # Apply formatting to all cells
    for col_idx, cell_value in enumerate(row):
        cell = ws_issues.cell(row=row_num, column=col_idx + 1)
        cell.font = CELL_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        if col_idx not in [3, 6]:  # Don't override severity and status colors
            fill_color = ALT_ROW_FILL if idx % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.fill = fill_color

# Set column widths
ws_issues.column_dimensions['A'].width = 12
ws_issues.column_dimensions['B'].width = 35
ws_issues.column_dimensions['C'].width = 25
ws_issues.column_dimensions['D'].width = 12
ws_issues.column_dimensions['E'].width = 35
ws_issues.column_dimensions['F'].width = 35
ws_issues.column_dimensions['G'].width = 18
ws_issues.row_dimensions[1].height = 25

# ===== METRICS SHEET =====
ws_metrics = wb.create_sheet('Metrics')
metrics_headers = ['Metric Category', 'Metric Name', 'Oracle Value', 'Snowflake Value', 'Change']
ws_metrics.append(metrics_headers)

# Format header row
for cell in ws_metrics[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

metrics_data = [
    ['Code Size', 'Total Functions', '3', '3', 'No change'],
    ['Code Size', 'Total Procedures', '4', '5 (one split due to overload)', '+1 additional'],
    ['Complexity', 'Components Ready', '7', '5', '-2 requiring work'],
    ['Effort', 'Estimated Hours', '0', '10.5', '+10.5 hours'],
    ['Data Types', 'Unique Types', '4', '4', 'Same count'],
    ['Oracle Features', 'Features to Convert', '5', '5', 'All mapped']
]

for idx, row in enumerate(metrics_data):
    ws_metrics.append(row)
    row_num = idx + 2
    
    # Apply formatting to all cells
    for col_idx, cell_value in enumerate(row):
        cell = ws_metrics.cell(row=row_num, column=col_idx + 1)
        cell.font = CELL_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        fill_color = ALT_ROW_FILL if idx % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.fill = fill_color

# Set column widths
ws_metrics.column_dimensions['A'].width = 18
ws_metrics.column_dimensions['B'].width = 25
ws_metrics.column_dimensions['C'].width = 15
ws_metrics.column_dimensions['D'].width = 25
ws_metrics.column_dimensions['E'].width = 20
ws_metrics.row_dimensions[1].height = 25

# ===== RECOMMENDATIONS SHEET =====
ws_recs = wb.create_sheet('Recommendations')
recs_headers = ['Phase', 'Timeline', 'Action', 'Priority', 'Owner']
ws_recs.append(recs_headers)

# Format header row
for cell in ws_recs[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

recs_data = [
    ['Phase 1: Preparation', 'Week 1', 'Define sequence strategy for auto-increment IDs', 'High', 'Database Architect'],
    ['Phase 1: Preparation', 'Week 1', 'Establish audit table schema and retention policies', 'High', 'Business Analyst'],
    ['Phase 1: Preparation', 'Week 1', 'Create Snowflake test environment', 'High', 'Infrastructure'],
    ['Phase 2: Core Migration', 'Week 2', 'Migrate get_customer and get_customer_name (simple reads)', 'High', 'Developer'],
    ['Phase 2: Core Migration', 'Week 2', 'Migrate delete_customer (simple deletes)', 'High', 'Developer'],
    ['Phase 2: Core Migration', 'Week 2', 'Migrate set_customer field-level variant', 'High', 'Developer'],
    ['Phase 2: Core Migration', 'Week 2', 'Migrate new_customer with sequence approach', 'High', 'Developer'],
    ['Phase 3: Complex Components', 'Week 3', 'Implement set_customer_row replacement procedure', 'Medium', 'Developer'],
    ['Phase 3: Complex Components', 'Week 3', 'Implement audit trail logic in purge_old_customers', 'High', 'Developer'],
    ['Phase 4: Validation', 'Week 4', 'Data migration validation and UAT', 'High', 'QA Team'],
    ['Phase 4: Validation', 'Week 4', 'Performance benchmarking and optimization', 'Medium', 'Database Performance']
]

for idx, row in enumerate(recs_data):
    ws_recs.append(row)
    row_num = idx + 2
    
    # Priority coloring
    priority_cell = ws_recs[f'D{row_num}']
    priority = row[3]
    if priority == 'High':
        priority_cell.fill = STATUS_HIGH
    elif priority == 'Medium':
        priority_cell.fill = STATUS_MEDIUM
    priority_cell.font = STATUS_FONT
    
    # Apply formatting to all cells
    for col_idx, cell_value in enumerate(row):
        cell = ws_recs.cell(row=row_num, column=col_idx + 1)
        cell.font = CELL_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        if col_idx != 3:  # Don't override priority color
            fill_color = ALT_ROW_FILL if idx % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.fill = fill_color

# Set column widths
ws_recs.column_dimensions['A'].width = 25
ws_recs.column_dimensions['B'].width = 12
ws_recs.column_dimensions['C'].width = 45
ws_recs.column_dimensions['D'].width = 12
ws_recs.column_dimensions['E'].width = 22
ws_recs.row_dimensions[1].height = 25

# Save workbook
output_file = r'E:\AGENT\.github\agents\OUTPUT\REPORTS\customer_pkg-status-report.xlsx'
wb.save(output_file)
print(f'Excel status report created: {output_file}')
print(f'Sheets created: {wb.sheetnames}')
print(f'\n✓ Professional formatting applied:')
print(f'  - Color-coded status indicators')
print(f'  - Professional header styling')
print(f'  - Alternating row colors')
print(f'  - Bordered cells')
print(f'  - Optimized column widths')
